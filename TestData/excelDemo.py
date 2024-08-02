import openpyxl

book = openpyxl.load_workbook("/Users/lap-082/Downloads/Untitled spreadsheet.xlsx")
sheet = book.active
cell = sheet.cell(row=2, column=2)
print(cell.value)

sheet.cell(row=4, column= 2).value = "musaumer@gmail.com"
print(sheet.cell(row=4, column= 2).value)

print(sheet.max_row)
print(sheet.max_column)